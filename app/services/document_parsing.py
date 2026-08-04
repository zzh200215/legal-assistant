"""#77/文档解析层（PDF/DOCX/MD/EXCEL/TXT/IMAGE + OCR + 分块/索引准备）

从 document_service.py 拆出（E-4），保持模块级函数签名与语义不变。
"""
import hashlib
import re
from pathlib import Path
from statistics import median

import pdfplumber
from docx import Document as DocxDocument
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook

from app.core.config import get_settings
from app.services.document_indexing import prepare_chunks_for_indexing
from app.services.storage_service import storage_service

settings = get_settings()

IMAGE_FILE_TYPES = {"png", ".png", "jpg", ".jpg", "jpeg", ".jpeg", "bmp", ".bmp", "webp", ".webp"}
VISION_SUPPORTED_FILE_TYPES = IMAGE_FILE_TYPES | {"pdf", ".pdf"}


class DocumentParsePermanentError(ValueError):
    pass


def _normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _normalize_conflict_label(value: str | None) -> str:
    """Keep fact matching deliberately conservative to avoid false conflict alerts."""
    text = (value or "").strip().lower()
    text = re.sub(r"[\s\W_]+", "", text)
    for token in ("日期", "时间", "期限", "截止", "金额", "数额", "费用", "负责人", "责任人", "负责", "相关", "事项"):
        text = text.replace(token, "")
    return text


def _facts_describe_same_subject(left: str | None, right: str | None) -> bool:
    left_label = _normalize_conflict_label(left)
    right_label = _normalize_conflict_label(right)
    if not left_label or not right_label:
        return False
    return left_label in right_label or right_label in left_label


def _build_section_path(parent_path: list[str] | None, title: str | None) -> list[str]:
    path = list(parent_path or [])
    normalized_title = (title or "").strip()
    if normalized_title:
        path.append(normalized_title)
    return path


def _unique_preserve_order(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        cleaned = (value or "").strip()
        if cleaned and cleaned not in result:
            result.append(cleaned)
    return result


def _humanize_visual_tag(tag: str) -> str:
    labels = {
        "ocr": "OCR识别",
        "visual": "视觉线索",
        "scanned_page": "扫描页",
        "page_visual": "页面视觉",
        "image_visual": "图片视觉",
        "table_visual": "表格视觉",
        "table_dense": "表格密集",
        "seal_present": "检测到公章",
        "stamp_present": "检测到印章",
        "signature_present": "检测到签字",
        "signed_page": "签署页",
        "attachment_like": "附件页",
        "image_like": "图像内容",
        "document_copy": "扫描件/复印件",
    }
    return labels.get(tag, tag.replace("_", " "))


def _detect_segment_type(text: str, section_title: str | None = None) -> str:
    normalized = (text or "").strip()
    if not normalized:
        return "empty"
    if re.match(r"^(\d+[\.\)]|[-*•])\s+", normalized):
        return "list"
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    if lines and all(("|" in line) or ("\t" in line) for line in lines[: min(3, len(lines))]):
        return "table"
    if len(lines) <= 2 and len(normalized) <= 40:
        return "heading"
    return "paragraph"


def _build_segment(
    *,
    text: str,
    page_number: int | None,
    section_title: str | None,
    section_path: list[str] | None = None,
    segment_type: str | None = None,
    visual_tags: list[str] | None = None,
    ocr_quality: float | None = None,
) -> dict:
    normalized_title = (section_title or "").strip() or "正文"
    normalized_text = _normalize_text(text)
    resolved_path = section_path or _build_section_path([], normalized_title)
    resolved_type = segment_type or _detect_segment_type(normalized_text, normalized_title)
    resolved_visual_tags = _derive_visual_tags(
        normalized_text,
        section_title=normalized_title,
        section_path=resolved_path,
        segment_type=resolved_type,
        existing_tags=visual_tags,
    )
    resolved_ocr_quality = ocr_quality
    if resolved_ocr_quality is None and resolved_type in {"page_ocr", "image_ocr"}:
        resolved_ocr_quality = round(_estimate_readable_ratio(normalized_text), 4)
    return {
        "text": normalized_text,
        "page_number": page_number,
        "section_title": normalized_title,
        "section_path": resolved_path,
        "segment_type": resolved_type,
        "visual_tags": resolved_visual_tags,
        "ocr_quality": resolved_ocr_quality,
        "visual_evidence": _extract_visual_evidence(normalized_text, visual_tags=resolved_visual_tags),
        "visual_region": None,
    }


def _load_ocr_dependencies():
    try:
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps  # type: ignore
        import pytesseract  # type: ignore

        return Image, ImageEnhance, ImageFilter, ImageOps, pytesseract
    except Exception:
        return None, None, None, None, None


def _safe_float(value, default: float | None = None) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _open_ocr_target(*, file_path: str | None = None, image=None):
    image_module, _, _, _, _ = _load_ocr_dependencies()
    if image_module is None:
        raise DocumentParsePermanentError("当前环境未启用 OCR，无法解析图片或扫描 PDF。")
    if image is not None:
        return image, None
    if not file_path:
        raise DocumentParsePermanentError("OCR 缺少输入文件。")
    opened_image = image_module.open(file_path)
    return opened_image, opened_image


def _collect_ocr_words(data: dict) -> list[dict]:
    texts = data.get("text") or []
    if not isinstance(texts, list):
        return []

    lefts = data.get("left") or []
    tops = data.get("top") or []
    widths = data.get("width") or []
    heights = data.get("height") or []
    confs = data.get("conf") or []
    words = []
    for index, raw_text in enumerate(texts):
        text = _normalize_text(str(raw_text or "").replace("\n", " "))
        if not text:
            continue
        conf = _safe_float(confs[index] if index < len(confs) else None, default=None)
        if conf is not None and conf < 0:
            continue
        left = _safe_int(lefts[index] if index < len(lefts) else 0)
        top = _safe_int(tops[index] if index < len(tops) else 0)
        width = max(1, _safe_int(widths[index] if index < len(widths) else 1))
        height = max(1, _safe_int(heights[index] if index < len(heights) else 1))
        words.append(
            {
                "text": text,
                "left": left,
                "top": top,
                "right": left + width,
                "width": width,
                "height": height,
                "conf": conf,
            }
        )
    words.sort(key=lambda item: (item["top"], item["left"]))
    return words


def _estimate_readable_ratio(text: str) -> float:
    normalized = _normalize_text(text or "")
    if not normalized:
        return 0.0
    meaningful = re.findall(r"[A-Za-z0-9\u4e00-\u9fff]", normalized)
    total = re.findall(r"\S", normalized)
    if not total:
        return 0.0
    return len(meaningful) / len(total)


def _estimate_table_density(text: str) -> float:
    normalized = _normalize_text(text or "")
    if not normalized:
        return 0.0
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    if not lines:
        return 0.0
    table_lines = sum(1 for line in lines if ("|" in line) or ("\t" in line))
    return table_lines / max(len(lines), 1)


def _derive_visual_tags(
    text: str,
    *,
    section_title: str | None = None,
    section_path: list[str] | None = None,
    segment_type: str | None = None,
    existing_tags: list[str] | None = None,
) -> list[str]:
    merged = " ".join(
        [
            text or "",
            section_title or "",
            " ".join(section_path or []),
            segment_type or "",
        ]
    )
    tags = list(existing_tags or [])
    if segment_type in {"page_ocr", "image_ocr"}:
        tags.extend(["ocr", "visual"])
    if segment_type == "page_ocr":
        tags.extend(["scanned_page", "page_visual"])
    if segment_type == "image_ocr":
        tags.extend(["image_visual"])
    if segment_type == "table" or _estimate_table_density(text) >= 0.35:
        tags.extend(["table_visual", "table_dense"])
    if re.search(r"(盖章|签章|公章|印章|骑缝章)", merged):
        tags.extend(["seal_present", "stamp_present"])
    if re.search(r"(签字|签署|签名|签约人|授权代表)", merged):
        tags.extend(["signature_present", "signed_page"])
    if re.search(r"(附件|附录|附件清单|附页)", merged):
        tags.extend(["attachment_like"])
    if re.search(r"(图片|截图|照片|扫描|影印)", merged):
        tags.extend(["image_like"])
    if re.search(r"(原件|复印件|扫描件)", merged):
        tags.extend(["document_copy"])
    return _unique_preserve_order(tags)


def _build_visual_summary(
    *,
    visual_tags: list[str] | None = None,
    segment_type: str | None = None,
    page_number: int | None = None,
    ocr_quality: float | None = None,
    section_title: str | None = None,
) -> str:
    tags = _unique_preserve_order(list(visual_tags or []))
    parts: list[str] = []
    if segment_type in {"page_ocr", "image_ocr"}:
        parts.append("来源: OCR")
    if page_number is not None:
        parts.append(f"页码: 第 {page_number} 页")
    if section_title:
        parts.append(f"标题: {section_title}")
    if tags:
        parts.append("视觉标签: " + "、".join(_humanize_visual_tag(tag) for tag in tags[:6]))
    if ocr_quality is not None:
        parts.append(f"OCR质量: {round(float(ocr_quality), 2):.2f}")
    if not parts:
        return ""
    return "[视觉摘要] " + "；".join(parts)


def _extract_visual_evidence(
    text: str,
    *,
    visual_tags: list[str] | None = None,
    max_lines: int = 2,
) -> str:
    normalized = _normalize_text(text or "")
    if not normalized:
        return ""
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    if not lines:
        return ""

    patterns = []
    tags = set(visual_tags or [])
    if tags.intersection({"seal_present", "stamp_present"}):
        patterns.append(r"(盖章|签章|公章|印章|骑缝章)")
    if tags.intersection({"signature_present", "signed_page"}):
        patterns.append(r"(签字|签署|签名|签约人|授权代表)")
    if "attachment_like" in tags:
        patterns.append(r"(附件|附录|附件清单|附页)")
    if "table_visual" in tags:
        patterns.append(r"(\||\t|表格|表头|金额|日期)")

    if not patterns:
        return ""

    matched_lines = []
    for line in lines:
        if any(re.search(pattern, line) for pattern in patterns):
            matched_lines.append(line)
        if len(matched_lines) >= max_lines:
            break
    return "\n".join(matched_lines)


def _resolve_vertical_region(top: int, min_top: int, max_bottom: int) -> str:
    span = max(1, max_bottom - min_top)
    ratio = (top - min_top) / span
    if ratio < 0.33:
        return "top"
    if ratio < 0.66:
        return "middle"
    return "bottom"


def _question_has_visual_hint(question: str) -> bool:
    return "补充视觉分析线索：" in (question or "")


def _extract_visual_evidence_with_positions(
    data: dict,
    *,
    visual_tags: list[str] | None = None,
    max_lines: int = 2,
) -> tuple[str, str | None]:
    words = _collect_ocr_words(data)
    rows = _group_ocr_words_into_rows(words)
    if not rows:
        return "", None

    patterns = []
    tags = set(visual_tags or [])
    if tags.intersection({"seal_present", "stamp_present"}):
        patterns.append(r"(盖章|签章|公章|印章|骑缝章)")
    if tags.intersection({"signature_present", "signed_page"}):
        patterns.append(r"(签字|签署|签名|签约人|授权代表)")
    if "attachment_like" in tags:
        patterns.append(r"(附件|附录|附件清单|附页)")
    if "table_visual" in tags:
        patterns.append(r"(\||\t|表格|表头|金额|日期)")
    if not patterns:
        return "", None

    min_top = min(word["top"] for word in words)
    max_bottom = max(word["top"] + word["height"] for word in words)
    matched_lines = []
    matched_regions = []
    for row in rows:
        line = " ".join(item["text"] for item in row).strip()
        if not line:
            continue
        if any(re.search(pattern, line) for pattern in patterns):
            matched_lines.append(line)
            matched_regions.append(_resolve_vertical_region(row[0]["top"], min_top, max_bottom))
        if len(matched_lines) >= max_lines:
            break
    if not matched_lines:
        return "", None
    region = matched_regions[0] if matched_regions else None
    return "\n".join(matched_lines), region


def _looks_like_low_quality_text(text: str) -> bool:
    normalized = _normalize_text(text or "")
    if not normalized:
        return True
    if len(normalized) < max(1, settings.OCR_MIN_TEXT_LENGTH):
        return True
    return _estimate_readable_ratio(normalized) < max(0.0, min(1.0, settings.OCR_MIN_READABLE_RATIO))


def _build_ocr_image_variants(image) -> list:
    if image is None:
        return []
    if not settings.OCR_ENABLE_IMAGE_PREPROCESS:
        return [image]

    _, image_enhance_module, image_filter_module, image_ops_module, _ = _load_ocr_dependencies()
    if image_enhance_module is None or image_filter_module is None or image_ops_module is None:
        return [image]

    variants = [image]
    try:
        grayscale = image.convert("L")
        autocontrast = image_ops_module.autocontrast(grayscale)
        variants.append(autocontrast)

        sharpened = autocontrast.filter(image_filter_module.SHARPEN)
        variants.append(sharpened)

        contrast_boost = image_enhance_module.Contrast(autocontrast).enhance(1.8)
        variants.append(contrast_boost)

        binary = contrast_boost.point(lambda value: 255 if value > 160 else 0)
        variants.append(binary)
    except Exception:
        return [image]

    unique_variants = []
    seen = set()
    for variant in variants:
        marker = (id(variant), getattr(variant, "mode", None), getattr(variant, "size", None))
        if marker in seen:
            continue
        seen.add(marker)
        unique_variants.append(variant)
    return unique_variants


def _group_ocr_words_into_rows(words: list[dict]) -> list[list[dict]]:
    if not words:
        return []
    row_threshold = max(8, int(median([item["height"] for item in words]) * 0.7))
    rows: list[list[dict]] = []
    current_row: list[dict] = []
    current_top = words[0]["top"]
    for word in words:
        if not current_row or abs(word["top"] - current_top) <= row_threshold:
            current_row.append(word)
            current_top = int(sum(item["top"] for item in current_row) / len(current_row))
            continue
        rows.append(sorted(current_row, key=lambda item: item["left"]))
        current_row = [word]
        current_top = word["top"]
    if current_row:
        rows.append(sorted(current_row, key=lambda item: item["left"]))
    return rows


def _split_ocr_row_into_cells(row: list[dict]) -> list[dict]:
    if not row:
        return []
    gap_threshold = max(14, int(median([item["width"] for item in row]) * 0.8))
    cells = []
    current_words = [row[0]]
    current_left = row[0]["left"]
    current_right = row[0]["right"]
    for word in row[1:]:
        gap = word["left"] - current_right
        if gap > gap_threshold:
            cells.append(
                {
                    "left": current_left,
                    "right": current_right,
                    "text": " ".join(item["text"] for item in current_words),
                }
            )
            current_words = [word]
            current_left = word["left"]
            current_right = word["right"]
        else:
            current_words.append(word)
            current_right = max(current_right, word["right"])
    cells.append(
        {
            "left": current_left,
            "right": current_right,
            "text": " ".join(item["text"] for item in current_words),
        }
    )
    return cells


def _cluster_column_positions(positions: list[int], tolerance: int) -> list[int]:
    if not positions:
        return []
    sorted_positions = sorted(positions)
    clusters: list[list[int]] = [[sorted_positions[0]]]
    for position in sorted_positions[1:]:
        if abs(position - int(sum(clusters[-1]) / len(clusters[-1]))) <= tolerance:
            clusters[-1].append(position)
        else:
            clusters.append([position])
    return [int(sum(cluster) / len(cluster)) for cluster in clusters]


def _render_table_from_ocr_data(data: dict) -> str | None:
    words = _collect_ocr_words(data)
    rows = _group_ocr_words_into_rows(words)
    if len(rows) < 2:
        return None

    row_cells = [_split_ocr_row_into_cells(row) for row in rows]
    candidate_rows = [cells for cells in row_cells if len(cells) >= 2]
    if len(candidate_rows) < 2:
        return None

    cell_widths = [max(1, cell["right"] - cell["left"]) for cells in candidate_rows for cell in cells]
    tolerance = max(24, int(median(cell_widths) * 0.6))
    anchors = _cluster_column_positions([cell["left"] for cells in candidate_rows for cell in cells], tolerance)
    if len(anchors) < 2:
        return None

    rendered_rows = []
    populated_rows = 0
    for cells in row_cells:
        values = [""] * len(anchors)
        populated = 0
        for cell in cells:
            anchor_index = min(range(len(anchors)), key=lambda idx: abs(cell["left"] - anchors[idx]))
            if values[anchor_index]:
                values[anchor_index] = f"{values[anchor_index]} {cell['text']}".strip()
            else:
                values[anchor_index] = cell["text"]
                populated += 1
        if populated >= 2:
            populated_rows += 1
        rendered_rows.append(values)

    if populated_rows < 2:
        return None

    non_empty_columns = [
        index for index in range(len(anchors))
        if any((row[index] or "").strip() for row in rendered_rows)
    ]
    if len(non_empty_columns) < 2:
        return None

    lines = []
    for row in rendered_rows:
        filtered = [(row[index] or "").strip() for index in non_empty_columns]
        if sum(1 for item in filtered if item) < 2:
            continue
        lines.append(f"| {' | '.join(filtered)} |")
    return _normalize_text("\n".join(lines)) if len(lines) >= 2 else None


def _ocr_image_to_table_text(*, file_path: str | None = None, image=None) -> str | None:
    image_module, _, _, _, pytesseract_module = _load_ocr_dependencies()
    if image_module is None or pytesseract_module is None:
        raise DocumentParsePermanentError("当前环境未启用 OCR，无法解析图片或扫描 PDF。")

    target_image, opened_image = _open_ocr_target(file_path=file_path, image=image)
    try:
        output_type = getattr(getattr(pytesseract_module, "Output", None), "DICT", None)
        for variant in _build_ocr_image_variants(target_image):
            for lang in ("chi_sim+eng", "eng", None):
                try:
                    kwargs = {"lang": lang} if lang else {}
                    if output_type is not None:
                        kwargs["output_type"] = output_type
                    data = pytesseract_module.image_to_data(variant, **kwargs)
                    if isinstance(data, dict):
                        table_text = _render_table_from_ocr_data(data)
                        if table_text:
                            return table_text
                except Exception:
                    continue
        return None
    finally:
        if opened_image is not None:
            opened_image.close()


def _ocr_image_to_text(*, file_path: str | None = None, image=None) -> str:
    image_module, _, _, _, pytesseract_module = _load_ocr_dependencies()
    if image_module is None or pytesseract_module is None:
        raise DocumentParsePermanentError("当前环境未启用 OCR，无法解析图片或扫描 PDF。")

    def run_with_lang(target, lang: str | None) -> str:
        kwargs = {"lang": lang} if lang else {}
        return pytesseract_module.image_to_string(target, **kwargs)

    tried_errors: list[str] = []
    best_candidate = ""
    target_image = image
    opened_image = None
    try:
        if target_image is None:
            target_image, opened_image = _open_ocr_target(file_path=file_path, image=image)

        for variant in _build_ocr_image_variants(target_image):
            for lang in ("chi_sim+eng", "eng", None):
                try:
                    raw_text = run_with_lang(variant, lang)
                    normalized = _normalize_text(raw_text)
                    if not normalized:
                        continue
                    if not _looks_like_low_quality_text(normalized):
                        return normalized
                    if len(normalized) > len(best_candidate):
                        best_candidate = normalized
                except Exception as exc:
                    tried_errors.append(str(exc))
        if best_candidate:
            return best_candidate
        raise DocumentParsePermanentError(
            "OCR 已启用，但未识别到可用文本。"
            if not tried_errors
            else f"OCR 识别失败：{tried_errors[-1]}"
        )
    finally:
        if opened_image is not None:
            opened_image.close()


def _ocr_image_to_text_with_layout(*, file_path: str | None = None, image=None, visual_tags: list[str] | None = None) -> tuple[str, str | None]:
    text = _ocr_image_to_text(file_path=file_path, image=image)
    _, _, _, _, pytesseract_module = _load_ocr_dependencies()
    if pytesseract_module is None:
        return text, None
    target_image = image
    opened_image = None
    try:
        if target_image is None:
            target_image, opened_image = _open_ocr_target(file_path=file_path, image=image)
        output_type = getattr(getattr(pytesseract_module, "Output", None), "DICT", None)
        if output_type is None:
            return text, None
        for variant in _build_ocr_image_variants(target_image):
            for lang in ("chi_sim+eng", "eng", None):
                try:
                    kwargs = {"lang": lang} if lang else {}
                    kwargs["output_type"] = output_type
                    data = pytesseract_module.image_to_data(variant, **kwargs)
                    if isinstance(data, dict):
                        evidence, region = _extract_visual_evidence_with_positions(
                            data,
                            visual_tags=visual_tags,
                        )
                        if evidence:
                            return text, region
                except Exception:
                    continue
        return text, None
    finally:
        if opened_image is not None:
            opened_image.close()


def _ocr_pdf_page(page) -> tuple[str, str, str | None]:
    try:
        page_image = page.to_image(resolution=max(120, settings.OCR_PDF_RENDER_DPI))
        pil_image = getattr(page_image, "original", None)
        if pil_image is None:
            raise DocumentParsePermanentError("当前 PDF 渲染能力不可用，无法对扫描页执行 OCR。")
        table_text = _ocr_image_to_table_text(image=pil_image)
        if table_text:
            return table_text, "table", None
        visual_tags = _derive_visual_tags("", segment_type="page_ocr")
        text, visual_region = _ocr_image_to_text_with_layout(image=pil_image, visual_tags=visual_tags)
        return text, "page_ocr", visual_region
    except DocumentParsePermanentError:
        raise
    except Exception as exc:
        raise DocumentParsePermanentError(f"扫描 PDF 页 OCR 失败：{exc}") from exc


def _extract_segments_from_pdf(file_path: str) -> list[dict]:
    segments = []
    ocr_attempted = False
    with pdfplumber.open(file_path) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            page_text = _normalize_text(page.extract_text() or "")
            segment_type = "page"
            visual_region = None
            if _looks_like_low_quality_text(page_text):
                ocr_attempted = True
                try:
                    page_text, segment_type, visual_region = _ocr_pdf_page(page)
                except DocumentParsePermanentError:
                    if _looks_like_low_quality_text(page_text):
                        page_text = ""
            if not page_text:
                continue
            segments.append(
                _build_segment(
                    text=page_text,
                    page_number=index,
                    section_title=f"第 {index} 页",
                    section_path=[f"第 {index} 页"],
                    segment_type=segment_type,
                    visual_tags=_derive_visual_tags(page_text, section_title=f"第 {index} 页", section_path=[f"第 {index} 页"], segment_type=segment_type),
                )
            )
            if visual_region and segments[-1].get("segment_type") in {"page_ocr", "image_ocr"}:
                segments[-1]["visual_region"] = visual_region
    if not segments and ocr_attempted:
        raise DocumentParsePermanentError("PDF 未提取到可读文本，且当前环境无法完成扫描页 OCR。")
    return segments


def _extract_segments_from_docx(file_path: str) -> list[dict]:
    doc = DocxDocument(file_path)
    segments = []
    current_title = "正文"
    current_path = [current_title]
    buffer: list[str] = []

    def flush() -> None:
        nonlocal buffer
        joined = _normalize_text("\n".join(buffer))
        if joined:
            segments.append(
                _build_segment(
                    text=joined,
                    page_number=None,
                    section_title=current_title,
                    section_path=current_path,
                )
            )
        buffer = []

    for paragraph in doc.paragraphs:
        text = (paragraph.text or "").strip()
        if not text:
            continue
        style_name = getattr(paragraph.style, "name", "") or ""
        if style_name.lower().startswith("heading"):
            flush()
            current_title = text
            current_path = _build_section_path([], current_title)
            continue
        buffer.append(text)
    flush()
    return segments


def _split_markdown_sections(md_text: str) -> list[dict]:
    lines = md_text.splitlines()
    sections: list[dict] = []
    current_title = "正文"
    current_path = [current_title]
    buffer: list[str] = []
    heading_stack: list[tuple[int, str]] = []

    def flush() -> None:
        nonlocal buffer
        joined = _normalize_text("\n".join(buffer))
        if joined:
            sections.append(
                _build_segment(
                    text=joined,
                    page_number=None,
                    section_title=current_title,
                    section_path=current_path,
                )
            )
        buffer = []

    for line in lines:
        stripped = line.strip()
        heading_match = re.match(r"^(#{1,6})\s+(.*)", stripped)
        if heading_match:
            flush()
            level = len(heading_match.group(1))
            title = heading_match.group(2).strip()
            heading_stack = [item for item in heading_stack if item[0] < level]
            heading_stack.append((level, title))
            current_title = title
            current_path = [item[1] for item in heading_stack]
            continue
        buffer.append(line)
    flush()
    return sections


def _extract_segments_from_markdown(file_path: str) -> list[dict]:
    with open(file_path, "r", encoding="utf-8") as f:
        md_text = f.read()
    return _split_markdown_sections(md_text)


def _extract_segments_from_excel(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, data_only=True)
    segments = []
    for index, sheet in enumerate(wb.worksheets, start=1):
        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                rows.append(row_text)
        joined = _normalize_text("\n".join(rows))
        if joined:
            segments.append(
                _build_segment(
                    text=joined,
                    page_number=index,
                    section_title=sheet.title,
                    section_path=[sheet.title],
                    segment_type="table",
                )
            )
    return segments


def _extract_segments_from_txt(file_path: str) -> list[dict]:
    with open(file_path, "r", encoding="utf-8") as f:
        text = _normalize_text(f.read())
    if not text:
        return []
    return [_build_segment(text=text, page_number=None, section_title="正文", section_path=["正文"])]


def _extract_segments_from_image(file_path: str) -> list[dict]:
    table_text = _ocr_image_to_table_text(file_path=file_path)
    if table_text:
        title = Path(file_path).name
        return [
            _build_segment(
                text=table_text,
                page_number=1,
                section_title=title,
                section_path=[title],
                segment_type="table",
            )
        ]

    title = Path(file_path).name
    visual_tags = _derive_visual_tags("", section_title=title, section_path=[title], segment_type="image_ocr")
    text, visual_region = _ocr_image_to_text_with_layout(
        file_path=file_path,
        visual_tags=visual_tags,
    )
    if not text:
        raise DocumentParsePermanentError("图片未识别到可用文本。")
    segment = _build_segment(
            text=text,
            page_number=1,
            section_title=title,
            section_path=[title],
            segment_type="image_ocr",
        )
    if visual_region:
        segment["visual_region"] = visual_region
    return [segment]


def _extract_segments(file_path: str, file_type: str) -> list[dict]:
    ext = file_type.lower()
    if ext in ("pdf", ".pdf"):
        return _extract_segments_from_pdf(file_path)
    if ext in ("docx", ".docx"):
        return _extract_segments_from_docx(file_path)
    if ext in ("xlsx", ".xlsx", "xls", ".xls"):
        return _extract_segments_from_excel(file_path)
    if ext in ("md", ".md", "markdown"):
        return _extract_segments_from_markdown(file_path)
    if ext in ("txt", ".txt"):
        return _extract_segments_from_txt(file_path)
    if ext in IMAGE_FILE_TYPES:
        return _extract_segments_from_image(file_path)
    raise DocumentParsePermanentError(f"Unsupported file type: {ext}")


def _extract_text(file_path: str, file_type: str) -> str:
    segments = _extract_segments(file_path, file_type)
    return _normalize_text("\n\n".join(segment["text"] for segment in segments))


def extract_file_text(file_path: str, file_type: str) -> str:
    return _extract_text(file_path, file_type)


def _file_to_data_url(file_path: str, file_type: str) -> str:
    mime_map = {
        "png": "image/png",
        ".png": "image/png",
        "jpg": "image/jpeg",
        ".jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        ".jpeg": "image/jpeg",
        "bmp": "image/bmp",
        ".bmp": "image/bmp",
        "webp": "image/webp",
        ".webp": "image/webp",
        "pdf": "application/pdf",
        ".pdf": "application/pdf",
    }
    normalized_type = (file_type or "").lower()
    mime_type = mime_map.get(normalized_type)
    if not mime_type:
        raise DocumentParsePermanentError("当前文件类型不支持视觉模型分析。")
    return storage_service.to_data_url(file_path, mime_type)


def _supports_visual_analysis(file_type: str) -> bool:
    return (file_type or "").lower() in VISION_SUPPORTED_FILE_TYPES


def _split_text(text_or_segments: str | list[dict], chunk_size: int = 800, chunk_overlap: int = 100) -> list[dict]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", "。", ".", " ", ""],
    )
    if isinstance(text_or_segments, str):
        segments = [_build_segment(text=text_or_segments, page_number=None, section_title="正文", section_path=["正文"])]
    else:
        segments = text_or_segments

    chunks = []
    chunk_index = 0
    for segment in segments:
        segment_text = _normalize_text(segment.get("text", ""))
        if not segment_text:
            continue
        for content in splitter.split_text(segment_text):
            normalized = _normalize_text(content)
            if not normalized:
                continue
            chunks.append(
                {
                    "chunk_index": chunk_index,
                    "content": normalized,
                    "page_number": segment.get("page_number"),
                    "section_title": segment.get("section_title"),
                    "section_path": segment.get("section_path") or [segment.get("section_title") or "正文"],
                    "segment_type": segment.get("segment_type") or _detect_segment_type(normalized, segment.get("section_title")),
                    "table_like": bool((segment.get("segment_type") == "table") or ("|" in normalized) or ("\t" in normalized)),
                    "visual_tags": _derive_visual_tags(
                        normalized,
                        section_title=segment.get("section_title"),
                        section_path=segment.get("section_path") or [segment.get("section_title") or "正文"],
                        segment_type=segment.get("segment_type") or _detect_segment_type(normalized, segment.get("section_title")),
                        existing_tags=segment.get("visual_tags") or [],
                    ),
                    "ocr_quality": segment.get("ocr_quality"),
                    "visual_evidence": segment.get("visual_evidence")
                    or _extract_visual_evidence(normalized, visual_tags=segment.get("visual_tags") or []),
                    "visual_region": segment.get("visual_region"),
                }
            )
            chunk_index += 1
    return chunks


def _fallback_summary_from_text(text: str, max_length: int) -> str:
    normalized = _normalize_text(text or "")
    if not normalized:
        return "文档内容为空，暂时无法生成摘要。"
    if len(normalized) <= max_length:
        return normalized
    return f"{normalized[:max_length].rstrip()}..."


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _prepare_chunks_for_indexing(document_id: int, chunks: list[dict]) -> list[dict]:
    return prepare_chunks_for_indexing(
        document_id,
        chunks,
        build_visual_summary=_build_visual_summary,
    )
